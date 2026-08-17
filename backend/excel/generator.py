"""Company 데이터를 엑셀 기업종합보고서로 변환한다 (openpyxl).

outputs/{기업명}_기업종합보고서.xlsx 로 저장 (목업 로그 문구와 동일한 파일명 규칙 —
다운로드용 파일명이라 사업자번호 대신 회사명을 써도 무방, backend/storage.py의
data/{사업자번호}.json 저장 키와는 별개).
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from ..models import Company

OUTPUT_DIR = Path(__file__).resolve().parents[2] / "outputs"

_TITLE_FONT = Font(size=16, bold=True)
_SECTION_FONT = Font(size=12, bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="1741A6")
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="E6EEFD")
_LABEL_FONT = Font(bold=True)

_YEARS = ("2023", "2024", "2025")


def _sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip() or "기업"


def _section_title(ws: Worksheet, row: int, text: str, span: int = 5) -> int:
    ws.cell(row=row, column=1, value=text).font = _SECTION_FONT
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = _SECTION_FILL
    return row + 1


def _kv_row(ws: Worksheet, row: int, label: str, value) -> int:
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    ws.cell(row=row, column=2, value=value if value is not None else "-")
    return row + 1


def _yearly_table(ws: Worksheet, row: int, rows: dict[str, dict[str, float | None]]) -> int:
    ws.cell(row=row, column=1, value="구분").font = _HEADER_FONT
    for i, year in enumerate(_YEARS, start=2):
        cell = ws.cell(row=row, column=i, value=year)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    row += 1
    for label, values in rows.items():
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        for i, year in enumerate(_YEARS, start=2):
            ws.cell(row=row, column=i, value=values.get(year))
        row += 1
    return row + 1


def _build_summary_sheet(ws: Worksheet, company: Company) -> None:
    ws.column_dimensions["A"].width = 20
    for col in "BCDE":
        ws.column_dimensions[col].width = 16

    ws.cell(row=1, column=1, value=f"기업종합보고서 — {company.company_name}").font = _TITLE_FONT
    row = 3

    row = _section_title(ws, row, "기본정보")
    for label, value in [
        ("기업명", company.company_name),
        ("사업자번호", company.business_no),
        ("대표자명", company.representative),
        ("주소", company.address),
        ("설립년월", company.founded_date),
        ("업종", company.industry_name),
        ("기업유형", company.company_type),
        ("기업규모", company.company_size),
    ]:
        row = _kv_row(ws, row, label, value)
    row += 1

    row = _section_title(ws, row, "등급")
    for label, value in [
        ("기업신용등급", company.credit_grade),
        ("EW등급", company.ew_grade),
        ("기업성장등급", company.growth_grade),
    ]:
        row = _kv_row(ws, row, label, value)
    row += 1

    row = _section_title(ws, row, "재무진단 (5축)")
    diag = company.diagnosis
    for label, value in [
        ("성장성", diag.growth),
        ("수익성", diag.profitability),
        ("재무구조", diag.financial_structure),
        ("부채상환능력", diag.debt_repayment),
        ("활동성", diag.activity),
    ]:
        row = _kv_row(ws, row, label, value)
    row += 1

    row = _section_title(ws, row, "업계 비교")
    row = _kv_row(ws, row, "업계 순위", (
        f"{company.industry_rank.sample_size}개사 중 {company.industry_rank.rank}위"
        if company.industry_rank.rank and company.industry_rank.sample_size else "-"
    ))
    row = _kv_row(ws, row, "보고서 기준", (
        f"평가일자 {company.evaluation_date} · 결산일자 {company.settlement_date}"
        if company.evaluation_date or company.settlement_date else "-"
    ))


def _build_financials_sheet(ws: Worksheet, company: Company) -> None:
    ws.column_dimensions["A"].width = 22
    for col in "BCD":
        ws.column_dimensions[col].width = 14

    row = 1
    ws.cell(row=row, column=1, value="재무상태표 요약 (백만원)").font = _TITLE_FONT
    row += 2
    row = _yearly_table(ws, row, company.balance_summary)

    ws.cell(row=row, column=1, value="손익계산서 요약 (백만원)").font = _TITLE_FONT
    row += 2
    row = _yearly_table(ws, row, company.income_summary)

    ws.cell(row=row, column=1, value="재무비율 (%)").font = _TITLE_FONT
    row += 2
    _yearly_table(ws, row, company.ratio_summary)


def _build_industry_sheet(ws: Worksheet, company: Company) -> None:
    ws.column_dimensions["A"].width = 16
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 14

    row = 1
    ws.cell(row=row, column=1, value="동종업계 내 경영규모 비교 (백만원)").font = _TITLE_FONT
    row += 2

    fields = ["총자산", "자본총계", "납입자본금", "매출액", "영업이익", "당기순이익"]
    ws.cell(row=row, column=1, value="구분").font = _HEADER_FONT
    for i, field in enumerate(fields, start=2):
        cell = ws.cell(row=row, column=i, value=field)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    row += 1
    for row_label in ["조회기업", "상위25%", "평균", "하위25%"]:
        values = company.peer_comparison.get(row_label, {})
        ws.cell(row=row, column=1, value=row_label).font = _LABEL_FONT
        for i, field in enumerate(fields, start=2):
            ws.cell(row=row, column=i, value=values.get(field))
        row += 1


def generate_excel(company: Company, output_dir: Path | None = None) -> Path:
    output_dir = output_dir or OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _build_summary_sheet(wb.active, company)
    wb.active.title = "요약"
    _build_financials_sheet(wb.create_sheet("재무제표"), company)
    _build_industry_sheet(wb.create_sheet("업계비교"), company)

    filename = f"{_sanitize_filename(company.company_name)}_기업종합보고서.xlsx"
    path = output_dir / filename
    wb.save(path)
    return path
