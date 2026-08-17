from datetime import datetime, timezone

from openpyxl import load_workbook

from backend.excel.generator import generate_excel
from backend.models import Company, DiagnosisRatings, IndustryRank


def _sample_company() -> Company:
    return Company(
        business_no="412-93-13689",
        company_name="옥산농원",
        representative="김종원",
        address="광주 나주시 봉황면 옥산유곡길 48-19",
        industry_name="양계업",
        credit_grade="bb+",
        ew_grade="정상",
        balance_summary={"자산총계": {"2023": 3944, "2024": 4000, "2025": 4367}},
        income_summary={"매출액": {"2023": 7154, "2024": 7241, "2025": 8307}},
        ratio_summary={"부채비율": {"2023": 770.81, "2024": 351.75, "2025": 313.37}},
        diagnosis=DiagnosisRatings(growth="양호", profitability="우수"),
        industry_rank=IndustryRank(rank=74, sample_size=79),
        peer_comparison={"조회기업": {"매출액": 7241}},
        parsed_at=datetime.now(timezone.utc),
    )


def test_generate_excel_creates_file_with_expected_sheets(tmp_path):
    path = generate_excel(_sample_company(), output_dir=tmp_path)
    assert path.exists()
    assert path.name == "옥산농원_기업종합보고서.xlsx"

    wb = load_workbook(path)
    assert wb.sheetnames == ["요약", "재무제표", "업계비교"]


def test_summary_sheet_contains_key_values(tmp_path):
    path = generate_excel(_sample_company(), output_dir=tmp_path)
    ws = load_workbook(path)["요약"]
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert "bb+" in values
    assert "79개사 중 74위" in values


def test_financials_sheet_contains_yearly_values(tmp_path):
    path = generate_excel(_sample_company(), output_dir=tmp_path)
    ws = load_workbook(path)["재무제표"]
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert 4367 in values  # 자산총계 2025
    assert 8307 in values  # 매출액 2025


def test_filename_sanitizes_illegal_windows_characters(tmp_path):
    company = _sample_company()
    company.company_name = '이상한:회사/이름*"?'
    path = generate_excel(company, output_dir=tmp_path)
    assert path.exists()
    assert not any(ch in path.name for ch in '\\/:*?"<>|')
