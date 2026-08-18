from datetime import datetime, timezone

import openpyxl
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend import category_list, storage
from backend.app import app
from backend.excel import generator as excel_generator
from backend.models import Company
from backend.routers import upload as upload_router


def _company(**overrides) -> Company:
    base = dict(
        business_no="412-93-13689",
        company_name="농업회사법인옥산농원",
        credit_grade="bb+",
        parsed_at=datetime.now(timezone.utc),
    )
    base.update(overrides)
    return Company(**base)


_HEADER_ROW = [
    "기업체명", "대표자명", "주소/연락처", "대표자 연령", "가업승계 여부", "형태",
    "업종", "reg_no", "설립년월", "상세업종", "기업등급", "주채권은행",
    "매출액", "영업이익", "당기순이익", "보험료", "배당유무", "법인경영체",
    "법인세감면", "기타",
]


def _write_category_xlsx(path):
    wb = openpyxl.Workbook()
    del wb["Sheet"]

    ws = wb.create_sheet("개인사업자(106)")
    ws.append(_HEADER_ROW)
    ws.append([
        "옥산농원", "김종원", "전남 나주시", None, None, "개인사업", "양계업",
        "412-93-13689", 19930101, None, "bb", "농협은행", "10억", "1억",
        "0.9억", None, None, None, None, None,
    ])
    ws.append([
        "미분석농장", "박대표", "전남 함평군", None, None, "개인사업", "양돈업",
        "111-11-11111", 20100101, None, "b+", "농협은행", "5억", "0.3억",
        "0.2억", None, None, None, None, None,
    ])

    ws2 = wb.create_sheet("일반법인(261)")
    ws2.append(_HEADER_ROW)
    ws2.append([
        "㈜테스트법인", "이대표", "전남 담양군", None, None, "일반법인", "식품제조업",
        "204311-0000000", 20050101, None, "bbb-", "기업은행", "50억", "3억",
        "2억", None, None, None, None, None,
    ])

    ws3 = wb.create_sheet("농업회사법인(219)")
    ws3.append(_HEADER_ROW)
    ws4 = wb.create_sheet("영농조합법인(108)")
    ws4.append(_HEADER_ROW)

    wb.save(path)


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "DATA_DIR", tmp_path)
    monkeypatch.setattr(upload_router, "UPLOADS_DIR", tmp_path / "uploads")
    monkeypatch.setattr(excel_generator, "OUTPUT_DIR", tmp_path / "outputs")

    xlsx_path = tmp_path / "추가메뉴.xlsx"
    _write_category_xlsx(xlsx_path)
    monkeypatch.setattr(category_list, "CATEGORY_LIST_PATH", xlsx_path)
    category_list._cache.clear()
    return TestClient(app)


def test_individual_category_matches_by_business_no(client):
    storage.save_company(_company())
    res = client.get("/api/category-list/individual")
    assert res.status_code == 200
    body = res.json()
    assert body["label"] == "개인사업자"
    assert body["total"] == 2
    by_name = {item["company_name"]: item for item in body["items"]}

    assert by_name["옥산농원"]["analyzed"] is True
    assert by_name["옥산농원"]["business_no"] == "412-93-13689"
    assert by_name["옥산농원"]["credit_grade"] == "bb+"  # 우리 파싱 데이터가 우선
    assert by_name["옥산농원"]["credit_grade_ref"] == "bb"  # 원본 시트값도 보존

    assert by_name["미분석농장"]["analyzed"] is False
    assert by_name["미분석농장"]["business_no"] is None
    assert by_name["미분석농장"]["revenue_ref"] == "5억"


def test_general_corp_category_matches_by_normalized_name(client):
    storage.save_company(_company(business_no="204311-0000000", company_name="테스트법인"))
    res = client.get("/api/category-list/general_corp")
    body = res.json()
    assert body["total"] == 1
    assert body["items"][0]["analyzed"] is True
    assert body["items"][0]["business_no"] == "204311-0000000"


def test_empty_sheet_returns_zero_total(client):
    res = client.get("/api/category-list/agri_corp")
    assert res.status_code == 200
    assert res.json()["total"] == 0


def test_unknown_category_returns_404(client):
    res = client.get("/api/category-list/does_not_exist")
    assert res.status_code == 404


def test_category_search(client):
    res = client.get("/api/category-list/individual", params={"q": "미분석"})
    body = res.json()
    assert body["total"] == 1
    assert body["items"][0]["company_name"] == "미분석농장"


def test_category_excel_export_preserves_all_columns(client):
    res = client.get("/api/category-list/individual/excel")
    assert res.status_code == 200
    import io
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws["A1"].value == "No."
    assert ws["B1"].value == "기업체명"
    assert ws["B2"].value == "옥산농원"
    assert ws["I2"].value == "412-93-13689"  # 사업자·법인등록번호 컬럼
