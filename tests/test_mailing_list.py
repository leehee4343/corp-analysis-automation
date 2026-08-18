from datetime import datetime, timezone

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend import storage
from backend.app import app
from backend.excel import generator as excel_generator
from backend.models import Company
from backend.routers import upload as upload_router


def _company(**overrides) -> Company:
    base = dict(
        business_no="412-93-13689",
        company_name="옥산농원",
        representative="김종원",
        address="전남나주시봉황면옥산유곡길48-19(옥산리)",
        postal_code="58235",
        parsed_at=datetime.now(timezone.utc),
    )
    base.update(overrides)
    return Company(**base)


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "DB_PATH", tmp_path / "test.db")
    monkeypatch.setattr(upload_router, "UPLOADS_DIR", tmp_path / "uploads")
    monkeypatch.setattr(excel_generator, "OUTPUT_DIR", tmp_path / "outputs")
    return TestClient(app)


def test_mailing_list_sorted_by_postal_code(client):
    storage.save_company(_company(
        business_no="303-81-54893", company_name="농업회사법인 동일농장",
        postal_code="58500", address="전남 어딘가",
    ))
    storage.save_company(_company())  # postal_code 58235, comes first

    res = client.get("/api/mailing-list")
    assert res.status_code == 200
    body = res.json()
    assert body["total"] == 2
    items = body["items"]
    assert items[0]["no"] == 1
    assert items[0]["company_name"] == "옥산농원"
    assert items[1]["no"] == 2
    assert items[1]["company_name"] == "농업회사법인 동일농장"


def test_mailing_list_pagination(client):
    for i in range(12):
        storage.save_company(_company(
            business_no=f"412-93-{13000 + i}", company_name=f"기업{i:02d}",
            postal_code=f"{10000 + i}",
        ))

    res = client.get("/api/mailing-list", params={"page": 1, "page_size": 10})
    body = res.json()
    assert body["total"] == 12
    assert len(body["items"]) == 10

    res2 = client.get("/api/mailing-list", params={"page": 2, "page_size": 10})
    assert len(res2.json()["items"]) == 2


def test_mailing_list_excel_matches_reference_format(client):
    storage.save_company(_company())
    res = client.get("/api/mailing-list/excel")
    assert res.status_code == 200
    assert res.headers["content-type"].startswith("application/vnd.openxmlformats")

    import io
    wb = load_workbook(io.BytesIO(res.content))
    assert wb.sheetnames == ["우편 발송용"]
    ws = wb.active
    assert [c.value for c in ws[1]] == ["No.", "우편번호", "주소", "상호명", "대표자 성명"]
    assert [c.value for c in ws[2]] == [1, "58235", "전남나주시봉황면옥산유곡길48-19(옥산리)", "옥산농원", "김종원"]
